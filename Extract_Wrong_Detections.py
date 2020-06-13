import openpyxl
wb = openpyxl.load_workbook('NUMBER_PLATE_LIST (Autosaved)1.xlsx')
ws=wb.active
count=0
i=0
first=ws['C']
second=ws['D']
a=[]
for(cell,cell_1) in zip(first,second):
    if str(cell.value) == str(cell_1.value).split()[-1]:
        count=count+1
    if str(cell.value) != str(cell_1.value).split()[-1]:
        count=count+1
        a.append(count)
        i=i+1

            

f=open("vehicle_list_day_footage.txt","r")
f1=open("out.txt","w")
count=0
i=0
x=0
for lines in f:
    count=count+1
    if(count == a[i]):
        print(lines)
        i=i+1
        f1.write(lines)
        continue



        

        
        
        
